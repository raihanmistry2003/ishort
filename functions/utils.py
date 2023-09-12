import pandas as pd
from openpyxl import load_workbook

def parse_csv_file(file_path, chunk_size=100):
    urls = []

    for chunk in pd.read_csv(file_path, chunksize=chunk_size):
        chunk_urls = chunk['Long Url'].dropna()
        urls.extend(chunk_urls)

    return urls

def parse_excel_file(file_path, chunk_size=10000):
    urls = []

    wb = load_workbook(file_path, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []

        for row in ws.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=data[0])
        sheet_urls = df['Long Url'].dropna()
        urls.extend(sheet_urls)

    return urls
